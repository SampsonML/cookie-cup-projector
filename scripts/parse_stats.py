"""Convert data/afl-fantasy-2026.xlsx to docs/data/players.json.

Header row is row 9 (1-indexed); player data starts at row 10. Blank cells are kept
as null in the output — projection logic on the frontend treats them as "did not
play," not zero.
"""
from __future__ import annotations
import json
import pathlib
from datetime import datetime, timezone

import openpyxl

REPO = pathlib.Path(__file__).resolve().parent.parent
XLSX_PATH = REPO / "data" / "afl-fantasy-2026.xlsx"
OUT_PATH = REPO / "docs" / "data" / "players.json"

HEADER_ROW = 9
DATA_START_ROW = 10


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _str(v):
    return str(v).strip() if v not in (None, "") else None


def _comp_block(row, start):
    """Each lower-league block is 4 cols: GMS, FP, MAX, 100+."""
    return {
        "games": _int(row[start]),
        "fp": _num(row[start + 1]),
        "max": _num(row[start + 2]),
        "c100": _int(row[start + 3]),
    }


def _has_signal(block):
    return any(v is not None for v in block.values())


def parse_row(row):
    name = _str(row[0])
    if not name:
        return None
    pos_raw = _str(row[6]) or ""
    positions = [p.strip() for p in pos_raw.split("/") if p.strip()]
    cd_id = _str(row[1])
    player_id = cd_id or f"{name}|{_str(row[3]) or ''}"

    record = {
        "id": player_id,
        "cd_id": cd_id,
        "name": name,
        "team": _str(row[3]),
        "salary": _int(row[4]),
        "owned": _num(row[5]),
        "positions": positions,
        "afl_2025": {
            "games": _int(row[7]),
            "fp": _num(row[8]),
            "max": _num(row[9]),
            "c100": _int(row[10]),
            "c120": _int(row[11]),
            "cba": _num(row[12]),
            "ppm": _num(row[13]),
            "reg": _num(row[14]),
            "l5": _num(row[15]),
            "fin": _num(row[16]),
        },
        "fp_2024": _num(row[17]),
        "fp_2023": _num(row[18]),
    }

    lower_leagues = {
        "vfl": _comp_block(row, 19),
        "wafl": _comp_block(row, 23),
        "sanfl": _comp_block(row, 27),
        "sanfl_u18": _comp_block(row, 31),
        "ctl": _comp_block(row, 35),
    }
    record["lower_leagues"] = {k: v for k, v in lower_leagues.items() if _has_signal(v)}
    return record


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["afl-fantasy-2026"] if "afl-fantasy-2026" in wb.sheetnames else wb.active

    players = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DATA_START_ROW:
            continue
        rec = parse_row(raw)
        if rec:
            players.append(rec)

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": str(XLSX_PATH.relative_to(REPO)),
        "count": len(players),
        "players": players,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, separators=(",", ":")))
    print(f"Wrote {len(players)} players to {OUT_PATH.relative_to(REPO)}")


if __name__ == "__main__":
    main()
